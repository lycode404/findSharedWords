import openpyxl
workbook = openpyxl.load_workbook('JISTable.xlsx')
sheet = workbook['Double Bytes']
count = 0
for code in range(0x0000,0xffff):
    try:
        jisbyte=bytes(divmod(code,0x100))
        jischr=str(jisbyte,encoding='shift-jis')
        gbkbyte=bytes(jischr,encoding='GB2312')
        if len(jischr) > 1:
            continue
        a,b = divmod(code,0x10)
        sheet[chr(66+b)+str(a+2)].fill = openpyxl.styles.PatternFill("solid", fgColor="1874CD")
        count += 1
    except:
        continue
print('Double bytes set has',count,'shared words.')
count = 0

sheet2 = workbook['Single Byte']

for code in range(0x00,0xff):
    try:
        jisbyte=bytes([code])
        jischr=str(jisbyte,encoding='shift-jis')
        gbkbyte=bytes(jischr,encoding='GB2312')
        if len(jischr) > 1:
            continue
        a,b = divmod(code,0x10)
        sheet2[chr(66+b)+str(a+2)].fill = openpyxl.styles.PatternFill("solid", fgColor="1874CD")
        count += 1
    except:
        continue
print('Single byte set has',count,'shared words.')
workbook.save('SharedWordsInJIS.xlsx')